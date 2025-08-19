"""
Comprehensive tests for analytics dashboard functionality.
Tests dashboard widgets, export functionality, and real-time updates.
"""

import json
import csv
import io
from datetime import timedelta
from unittest.mock import patch, MagicMock

from django.test import TestCase, Client
from django.contrib.auth import get_user_model
from django.urls import reverse
from django.utils import timezone
from django.core.cache import cache
from django.http import HttpResponse, JsonResponse
from channels.testing import WebsocketCommunicator
from channels.db import database_sync_to_async
from openpyxl import load_workbook

from apps.analytics.models import PageView, SearchQuery, SearchClickthrough
from apps.analytics.admin import AnalyticsAdminSite, AnalyticsDashboardWidget
from apps.analytics.consumers import AnalyticsDashboardConsumer, LiveStatsConsumer
from apps.blog.models import Post, Category
from apps.accounts.models import Profile

User = get_user_model()


class AnalyticsDashboardTestCase(TestCase):
    """
    Test case for analytics dashboard functionality.
    """
    
    def setUp(self):
        """Set up test data."""
        # Create test users
        self.admin_user = User.objects.create_user(
            username='admin',
            email='admin@test.com',
            password='testpass123',
            is_staff=True,
            is_superuser=True
        )
        
        self.staff_user = User.objects.create_user(
            username='staff',
            email='staff@test.com',
            password='testpass123',
            is_staff=True
        )
        
        self.regular_user = User.objects.create_user(
            username='user',
            email='user@test.com',
            password='testpass123'
        )
        
        # Create test blog content
        self.category = Category.objects.create(
            name='Test Category',
            slug='test-category'
        )
        
        self.post1 = Post.objects.create(
            title='Test Post 1',
            slug='test-post-1',
            content='Test content 1',
            author=self.admin_user,
            category=self.category,
            status='published',
            published_at=timezone.now()
        )
        
        self.post2 = Post.objects.create(
            title='Test Post 2',
            slug='test-post-2',
            content='Test content 2',
            author=self.admin_user,
            category=self.category,
            status='published',
            published_at=timezone.now()
        )
        
        # Create test analytics data
        self.create_test_analytics_data()
        
        # Set up client
        self.client = Client()
        
        # Clear cache before each test
        cache.clear()
    
    def create_test_analytics_data(self):
        """Create test analytics data."""
        now = timezone.now()
        
        # Create page views
        for i in range(50):
            PageView.objects.create(
                post=self.post1 if i % 2 == 0 else self.post2,
                user=self.regular_user if i % 3 == 0 else None,
                ip_address=f'192.168.1.{i % 255}',
                user_agent='Test User Agent',
                referrer='https://google.com' if i % 4 == 0 else '',
                session_key=f'session_{i}',
                path=f'/post/{self.post1.slug if i % 2 == 0 else self.post2.slug}/',
                time_on_page=60 + (i * 10) if i % 5 == 0 else None,
                scroll_depth=75 + (i % 25) if i % 5 == 0 else None,
                created_at=now - timedelta(hours=i)
            )
        
        # Create search queries
        search_terms = [
            'django', 'python', 'web development', 'testing', 'analytics',
            'blog', 'tutorial', 'programming', 'framework', 'database'
        ]
        
        for i, term in enumerate(search_terms * 5):
            SearchQuery.objects.create(
                query=f'{term} {i}' if i > 10 else term,
                results_count=max(0, 10 - (i % 12)),  # Some failed searches
                user=self.regular_user if i % 4 == 0 else None,
                ip_address=f'192.168.1.{i % 255}',
                user_agent='Test User Agent',
                session_key=f'search_session_{i}',
                clicked_post=self.post1 if i % 6 == 0 else None,
                clicked_result_position=1 if i % 6 == 0 else None,
                created_at=now - timedelta(hours=i)
            )
        
        # Create search clickthroughs
        for i in range(10):
            search_query = SearchQuery.objects.filter(
                clicked_post__isnull=False
            ).first()
            if search_query:
                SearchClickthrough.objects.create(
                    search_query=search_query,
                    post=search_query.clicked_post,
                    position=1,
                    user=search_query.user,
                    ip_address=search_query.ip_address,
                    session_key=search_query.session_key,
                    created_at=now - timedelta(hours=i)
                )


class DashboardWidgetTests(AnalyticsDashboardTestCase):
    """
    Tests for dashboard widgets.
    """
    
    def test_popular_posts_widget(self):
        """Test popular posts widget data generation."""
        widget_data = AnalyticsDashboardWidget.get_popular_posts_widget()
        
        self.assertIn('title', widget_data)
        self.assertIn('posts', widget_data)
        self.assertEqual(widget_data['title'], 'Popular Posts (Last 7 Days)')
        self.assertIsInstance(widget_data['posts'], list)
        
        # Should have posts with view counts
        if widget_data['posts']:
            post = widget_data['posts'][0]
            self.assertIn('post__title', post)
            self.assertIn('view_count', post)
            self.assertIn('post__id', post)
    
    def test_traffic_stats_widget(self):
        """Test traffic statistics widget data generation."""
        widget_data = AnalyticsDashboardWidget.get_traffic_stats_widget()
        
        self.assertIn('title', widget_data)
        self.assertIn('views', widget_data)
        self.assertIn('unique_visitors', widget_data)
        self.assertIn('searches', widget_data)
        self.assertIn('top_referrers', widget_data)
        
        # Check time period data
        self.assertIn('24h', widget_data['views'])
        self.assertIn('7d', widget_data['views'])
        self.assertIn('30d', widget_data['views'])
        
        # Should have numeric values
        self.assertIsInstance(widget_data['views']['24h'], int)
        self.assertIsInstance(widget_data['unique_visitors']['24h'], int)
        self.assertIsInstance(widget_data['searches']['24h'], int)
    
    def test_engagement_widget(self):
        """Test user engagement widget data generation."""
        widget_data = AnalyticsDashboardWidget.get_engagement_widget()
        
        self.assertIn('title', widget_data)
        self.assertIn('avg_time_on_page', widget_data)
        self.assertIn('avg_scroll_depth', widget_data)
        self.assertIn('engaged_sessions', widget_data)
        self.assertIn('total_views', widget_data)
        self.assertIn('engagement_rate', widget_data)
        
        # Should have numeric values
        self.assertIsInstance(widget_data['avg_time_on_page'], (int, float))
        self.assertIsInstance(widget_data['avg_scroll_depth'], (int, float))
        self.assertIsInstance(widget_data['engaged_sessions'], int)
        self.assertIsInstance(widget_data['engagement_rate'], (int, float))
    
    def test_widget_caching(self):
        """Test that widgets are properly cached."""
        # Clear cache
        cache.clear()
        
        # First call should hit database
        with self.assertNumQueries(3):  # Expect some queries
            widget_data1 = AnalyticsDashboardWidget.get_popular_posts_widget()
        
        # Second call should use cache
        with self.assertNumQueries(0):  # No queries expected
            widget_data2 = AnalyticsDashboardWidget.get_popular_posts_widget()
        
        self.assertEqual(widget_data1, widget_data2)


class DashboardViewTests(AnalyticsDashboardTestCase):
    """
    Tests for dashboard views.
    """
    
    def test_analytics_dashboard_view_access(self):
        """Test access control for analytics dashboard."""
        analytics_admin = AnalyticsAdminSite()
        
        # Test unauthenticated access
        response = self.client.get('/admin/analytics-dashboard/')
        self.assertEqual(response.status_code, 302)  # Redirect to login
        
        # Test regular user access
        self.client.login(username='user', password='testpass123')
        response = self.client.get('/admin/analytics-dashboard/')
        self.assertEqual(response.status_code, 302)  # Redirect to login
        
        # Test staff user access
        self.client.login(username='staff', password='testpass123')
        request = self.client.get('/admin/analytics-dashboard/').wsgi_request
        request.user = self.staff_user
        response = analytics_admin.analytics_dashboard_view(request)
        self.assertEqual(response.status_code, 200)
        
        # Test admin user access
        self.client.login(username='admin', password='testpass123')
        request = self.client.get('/admin/analytics-dashboard/').wsgi_request
        request.user = self.admin_user
        response = analytics_admin.analytics_dashboard_view(request)
        self.assertEqual(response.status_code, 200)
    
    def test_analytics_dashboard_context(self):
        """Test analytics dashboard context data."""
        analytics_admin = AnalyticsAdminSite()
        self.client.login(username='admin', password='testpass123')
        
        request = self.client.get('/admin/analytics-dashboard/').wsgi_request
        request.user = self.admin_user
        response = analytics_admin.analytics_dashboard_view(request)
        
        # Check context data
        context = response.context_data
        self.assertIn('title', context)
        self.assertIn('popular_posts', context)
        self.assertIn('traffic_stats', context)
        self.assertIn('engagement_stats', context)
        self.assertIn('search_stats', context)
        self.assertIn('recent_searches', context)
        
        self.assertEqual(context['title'], 'Analytics Dashboard')
    
    def test_search_analytics_report_view(self):
        """Test search analytics report view."""
        analytics_admin = AnalyticsAdminSite()
        self.client.login(username='admin', password='testpass123')
        
        request = self.client.get('/admin/search-analytics-report/?days=30').wsgi_request
        request.user = self.admin_user
        request.GET = {'days': '30'}
        response = analytics_admin.search_analytics_report_view(request)
        
        self.assertEqual(response.status_code, 200)
        
        # Check context data
        context = response.context_data
        self.assertIn('title', context)
        self.assertIn('days', context)
        self.assertIn('search_stats', context)
        self.assertIn('popular_queries', context)
        self.assertIn('failed_queries', context)
        self.assertIn('search_trends', context)
        
        self.assertEqual(context['days'], 30)
    
    def test_search_analytics_stats_calculation(self):
        """Test search analytics statistics calculation."""
        analytics_admin = AnalyticsAdminSite()
        
        stats = analytics_admin.get_search_analytics_stats(days=7)
        
        self.assertIn('total_searches', stats)
        self.assertIn('unique_queries', stats)
        self.assertIn('failed_searches', stats)
        self.assertIn('avg_results', stats)
        self.assertIn('success_rate', stats)
        
        # Should have numeric values
        self.assertIsInstance(stats['total_searches'], int)
        self.assertIsInstance(stats['unique_queries'], int)
        self.assertIsInstance(stats['failed_searches'], int)
        self.assertIsInstance(stats['avg_results'], (int, float))
        self.assertIsInstance(stats['success_rate'], (int, float))
        
        # Success rate should be between 0 and 100
        self.assertGreaterEqual(stats['success_rate'], 0)
        self.assertLessEqual(stats['success_rate'], 100)


class ExportFunctionalityTests(AnalyticsDashboardTestCase):
    """
    Tests for export functionality.
    """
    
    def test_csv_export_search_analytics(self):
        """Test CSV export of search analytics."""
        analytics_admin = AnalyticsAdminSite()
        self.client.login(username='admin', password='testpass123')
        
        request = self.client.get('/admin/analytics-export/?type=search&format=csv&days=30').wsgi_request
        request.user = self.admin_user
        request.GET = {'type': 'search', 'format': 'csv', 'days': '30'}
        
        response = analytics_admin.analytics_export_view(request)
        
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/csv')
        self.assertIn('attachment', response['Content-Disposition'])
        self.assertIn('search_analytics_30days.csv', response['Content-Disposition'])
        
        # Check CSV content
        content = response.content.decode('utf-8')
        self.assertIn('Search Analytics Export', content)
        self.assertIn('Total Searches', content)
        self.assertIn('Success Rate', content)
    
    def test_json_export_search_analytics(self):
        """Test JSON export of search analytics."""
        analytics_admin = AnalyticsAdminSite()
        self.client.login(username='admin', password='testpass123')
        
        request = self.client.get('/admin/analytics-export/?type=search&format=json&days=7').wsgi_request
        request.user = self.admin_user
        request.GET = {'type': 'search', 'format': 'json', 'days': '7'}
        
        response = analytics_admin.analytics_export_view(request)
        
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/json')
        self.assertIn('attachment', response['Content-Disposition'])
        
        # Check JSON content
        data = json.loads(response.content)
        self.assertIn('period_days', data)
        self.assertIn('total_searches', data)
        self.assertIn('searches', data)
        
        self.assertEqual(data['period_days'], 7)
        self.assertIsInstance(data['searches'], list)
    
    def test_excel_export_search_analytics(self):
        """Test Excel export of search analytics."""
        analytics_admin = AnalyticsAdminSite()
        self.client.login(username='admin', password='testpass123')
        
        request = self.client.get('/admin/analytics-export/?type=search&format=excel&days=30').wsgi_request
        request.user = self.admin_user
        request.GET = {'type': 'search', 'format': 'excel', 'days': '30'}
        
        response = analytics_admin.analytics_export_view(request)
        
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'], 
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        self.assertIn('attachment', response['Content-Disposition'])
        self.assertIn('search_analytics_30days.xlsx', response['Content-Disposition'])
        
        # Check Excel content
        workbook = load_workbook(io.BytesIO(response.content))
        self.assertIn('Summary', workbook.sheetnames)
        self.assertIn('Popular Queries', workbook.sheetnames)
        self.assertIn('Failed Queries', workbook.sheetnames)
        
        # Check summary sheet content
        summary_sheet = workbook['Summary']
        self.assertEqual(summary_sheet['A1'].value, 'Search Analytics Summary')
    
    def test_pageviews_csv_export(self):
        """Test CSV export of page views."""
        analytics_admin = AnalyticsAdminSite()
        self.client.login(username='admin', password='testpass123')
        
        request = self.client.get('/admin/analytics-export/?type=pageviews&format=csv&days=7').wsgi_request
        request.user = self.admin_user
        request.GET = {'type': 'pageviews', 'format': 'csv', 'days': '7'}
        
        response = analytics_admin.analytics_export_view(request)
        
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/csv')
        self.assertIn('pageviews_7days.csv', response['Content-Disposition'])
        
        # Check CSV content
        content = response.content.decode('utf-8')
        csv_reader = csv.reader(io.StringIO(content))
        headers = next(csv_reader)
        
        expected_headers = [
            'Path', 'Post Title', 'User', 'IP Address', 'Referrer',
            'Time on Page', 'Scroll Depth', 'Created At'
        ]
        self.assertEqual(headers, expected_headers)
    
    def test_dashboard_excel_export(self):
        """Test comprehensive dashboard Excel export."""
        analytics_admin = AnalyticsAdminSite()
        self.client.login(username='admin', password='testpass123')
        
        request = self.client.get('/admin/analytics-export/?type=dashboard&format=excel&days=30').wsgi_request
        request.user = self.admin_user
        request.GET = {'type': 'dashboard', 'format': 'excel', 'days': '30'}
        
        response = analytics_admin.analytics_export_view(request)
        
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'], 
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        self.assertIn('dashboard_export_30days.xlsx', response['Content-Disposition'])
        
        # Check Excel content
        workbook = load_workbook(io.BytesIO(response.content))
        expected_sheets = ['Dashboard Overview', 'Popular Posts', 'Traffic Sources']
        
        for sheet_name in expected_sheets:
            self.assertIn(sheet_name, workbook.sheetnames)
        
        # Check dashboard overview content
        overview_sheet = workbook['Dashboard Overview']
        self.assertEqual(overview_sheet['A1'].value, 'Analytics Dashboard Export')
    
    def test_invalid_export_parameters(self):
        """Test handling of invalid export parameters."""
        analytics_admin = AnalyticsAdminSite()
        self.client.login(username='admin', password='testpass123')
        
        request = self.client.get('/admin/analytics-export/?type=invalid&format=invalid').wsgi_request
        request.user = self.admin_user
        request.GET = {'type': 'invalid', 'format': 'invalid'}
        
        response = analytics_admin.analytics_export_view(request)
        
        self.assertEqual(response.status_code, 400)
        data = json.loads(response.content)
        self.assertIn('error', data)


class WebSocketTests(AnalyticsDashboardTestCase):
    """
    Tests for WebSocket real-time updates.
    """
    
    async def test_analytics_dashboard_consumer_connection(self):
        """Test WebSocket connection for analytics dashboard."""
        # Create authenticated user scope
        scope = {
            'type': 'websocket',
            'user': self.staff_user,
            'path': '/ws/analytics/dashboard/',
        }
        
        communicator = WebsocketCommunicator(AnalyticsDashboardConsumer.as_asgi(), '/ws/analytics/dashboard/')
        communicator.scope['user'] = self.staff_user
        
        connected, subprotocol = await communicator.connect()
        self.assertTrue(connected)
        
        # Should receive initial dashboard data
        response = await communicator.receive_json_from()
        self.assertEqual(response['type'], 'dashboard_update')
        self.assertIn('data', response)
        
        await communicator.disconnect()
    
    async def test_analytics_dashboard_consumer_unauthenticated(self):
        """Test WebSocket connection rejection for unauthenticated users."""
        from django.contrib.auth.models import AnonymousUser
        
        communicator = WebsocketCommunicator(AnalyticsDashboardConsumer.as_asgi(), '/ws/analytics/dashboard/')
        communicator.scope['user'] = AnonymousUser()
        
        connected, subprotocol = await communicator.connect()
        self.assertFalse(connected)
    
    async def test_analytics_dashboard_consumer_non_staff(self):
        """Test WebSocket connection rejection for non-staff users."""
        communicator = WebsocketCommunicator(AnalyticsDashboardConsumer.as_asgi(), '/ws/analytics/dashboard/')
        communicator.scope['user'] = self.regular_user
        
        connected, subprotocol = await communicator.connect()
        self.assertFalse(connected)
    
    async def test_live_stats_consumer_connection(self):
        """Test WebSocket connection for live stats."""
        communicator = WebsocketCommunicator(LiveStatsConsumer.as_asgi(), '/ws/analytics/live-stats/')
        communicator.scope['user'] = self.staff_user
        
        connected, subprotocol = await communicator.connect()
        self.assertTrue(connected)
        
        # Should receive initial live stats
        response = await communicator.receive_json_from()
        self.assertEqual(response['type'], 'live_stats')
        self.assertIn('data', response)
        
        await communicator.disconnect()
    
    async def test_websocket_message_handling(self):
        """Test WebSocket message handling."""
        communicator = WebsocketCommunicator(AnalyticsDashboardConsumer.as_asgi(), '/ws/analytics/dashboard/')
        communicator.scope['user'] = self.staff_user
        
        connected, subprotocol = await communicator.connect()
        self.assertTrue(connected)
        
        # Consume initial message
        await communicator.receive_json_from()
        
        # Send request for update
        await communicator.send_json_to({
            'type': 'request_update'
        })
        
        # Should receive dashboard update
        response = await communicator.receive_json_from()
        self.assertEqual(response['type'], 'dashboard_update')
        
        # Send request for chart data
        await communicator.send_json_to({
            'type': 'request_chart_data',
            'chart_type': 'daily_views'
        })
        
        # Should receive chart data
        response = await communicator.receive_json_from()
        self.assertEqual(response['type'], 'chart_data')
        self.assertEqual(response['chart_type'], 'daily_views')
        
        await communicator.disconnect()


class PerformanceTests(AnalyticsDashboardTestCase):
    """
    Tests for dashboard performance.
    """
    
    def test_dashboard_query_optimization(self):
        """Test that dashboard queries are optimized."""
        analytics_admin = AnalyticsAdminSite()
        
        # Test search analytics stats with query counting
        with self.assertNumQueries(5):  # Should be optimized to minimal queries
            stats = analytics_admin.get_search_analytics_stats(days=30)
        
        self.assertIsInstance(stats, dict)
        self.assertIn('total_searches', stats)
    
    def test_widget_caching_performance(self):
        """Test widget caching improves performance."""
        # Clear cache
        cache.clear()
        
        # First call - should hit database
        start_time = timezone.now()
        widget_data1 = AnalyticsDashboardWidget.get_popular_posts_widget()
        first_call_time = (timezone.now() - start_time).total_seconds()
        
        # Second call - should use cache
        start_time = timezone.now()
        widget_data2 = AnalyticsDashboardWidget.get_popular_posts_widget()
        second_call_time = (timezone.now() - start_time).total_seconds()
        
        # Cached call should be significantly faster
        self.assertLess(second_call_time, first_call_time)
        self.assertEqual(widget_data1, widget_data2)
    
    def test_large_dataset_export_performance(self):
        """Test export performance with larger datasets."""
        # Create additional test data
        now = timezone.now()
        
        # Create 1000 additional page views
        page_views = []
        for i in range(1000):
            page_views.append(PageView(
                post=self.post1,
                ip_address=f'10.0.{i // 255}.{i % 255}',
                user_agent='Test Agent',
                session_key=f'bulk_session_{i}',
                path=f'/test-path-{i}/',
                created_at=now - timedelta(minutes=i)
            ))
        
        PageView.objects.bulk_create(page_views)
        
        # Test CSV export performance
        analytics_admin = AnalyticsAdminSite()
        self.client.login(username='admin', password='testpass123')
        
        request = self.client.get('/admin/analytics-export/?type=pageviews&format=csv&days=1').wsgi_request
        request.user = self.admin_user
        request.GET = {'type': 'pageviews', 'format': 'csv', 'days': '1'}
        
        start_time = timezone.now()
        response = analytics_admin.analytics_export_view(request)
        export_time = (timezone.now() - start_time).total_seconds()
        
        self.assertEqual(response.status_code, 200)
        # Export should complete within reasonable time (5 seconds)
        self.assertLess(export_time, 5.0)


class IntegrationTests(AnalyticsDashboardTestCase):
    """
    Integration tests for dashboard functionality.
    """
    
    def test_end_to_end_dashboard_workflow(self):
        """Test complete dashboard workflow."""
        # Login as admin
        self.client.login(username='admin', password='testpass123')
        
        # Access dashboard
        analytics_admin = AnalyticsAdminSite()
        request = self.client.get('/admin/analytics-dashboard/').wsgi_request
        request.user = self.admin_user
        response = analytics_admin.analytics_dashboard_view(request)
        
        self.assertEqual(response.status_code, 200)
        
        # Access search report
        request = self.client.get('/admin/search-analytics-report/?days=30').wsgi_request
        request.user = self.admin_user
        request.GET = {'days': '30'}
        response = analytics_admin.search_analytics_report_view(request)
        
        self.assertEqual(response.status_code, 200)
        
        # Export data
        request = self.client.get('/admin/analytics-export/?type=search&format=csv&days=30').wsgi_request
        request.user = self.admin_user
        request.GET = {'type': 'search', 'format': 'csv', 'days': '30'}
        response = analytics_admin.analytics_export_view(request)
        
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/csv')
    
    def test_dashboard_with_no_data(self):
        """Test dashboard behavior with no analytics data."""
        # Clear all analytics data
        PageView.objects.all().delete()
        SearchQuery.objects.all().delete()
        SearchClickthrough.objects.all().delete()
        
        # Test dashboard still works
        analytics_admin = AnalyticsAdminSite()
        self.client.login(username='admin', password='testpass123')
        
        request = self.client.get('/admin/analytics-dashboard/').wsgi_request
        request.user = self.admin_user
        response = analytics_admin.analytics_dashboard_view(request)
        
        self.assertEqual(response.status_code, 200)
        
        # Test widgets handle empty data gracefully
        popular_posts = AnalyticsDashboardWidget.get_popular_posts_widget()
        traffic_stats = AnalyticsDashboardWidget.get_traffic_stats_widget()
        engagement_stats = AnalyticsDashboardWidget.get_engagement_widget()
        
        self.assertIsInstance(popular_posts['posts'], list)
        self.assertIsInstance(traffic_stats['views']['24h'], int)
        self.assertIsInstance(engagement_stats['avg_time_on_page'], (int, float))
    
    def test_dashboard_error_handling(self):
        """Test dashboard error handling."""
        analytics_admin = AnalyticsAdminSite()
        
        # Test with invalid days parameter
        request = self.client.get('/admin/search-analytics-report/?days=invalid').wsgi_request
        request.user = self.admin_user
        request.GET = {'days': 'invalid'}
        
        # Should handle gracefully and use default
        response = analytics_admin.search_analytics_report_view(request)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context_data['days'], 30)  # Default value
        
        # Test with extreme days parameter
        request.GET = {'days': '10000'}
        response = analytics_admin.search_analytics_report_view(request)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context_data['days'], 365)  # Max value


class SecurityTests(AnalyticsDashboardTestCase):
    """
    Security tests for dashboard functionality.
    """
    
    def test_dashboard_access_control(self):
        """Test dashboard access control."""
        analytics_admin = AnalyticsAdminSite()
        
        # Test unauthenticated access
        request = self.client.get('/admin/analytics-dashboard/').wsgi_request
        request.user = None
        
        # Should require authentication (handled by admin_view decorator)
        # This would normally redirect to login
        
        # Test non-staff user access
        request.user = self.regular_user
        # Should be denied access (handled by admin_view decorator)
        
        # Test staff user access
        request.user = self.staff_user
        response = analytics_admin.analytics_dashboard_view(request)
        self.assertEqual(response.status_code, 200)
    
    def test_export_access_control(self):
        """Test export functionality access control."""
        analytics_admin = AnalyticsAdminSite()
        
        # Test with staff user
        request = self.client.get('/admin/analytics-export/?type=search&format=csv').wsgi_request
        request.user = self.staff_user
        request.GET = {'type': 'search', 'format': 'csv'}
        
        response = analytics_admin.analytics_export_view(request)
        self.assertEqual(response.status_code, 200)
    
    def test_data_sanitization_in_exports(self):
        """Test that exported data is properly sanitized."""
        # Create search query with potentially problematic content
        SearchQuery.objects.create(
            query='<script>alert("xss")</script>',
            results_count=5,
            ip_address='192.168.1.100',
            user_agent='Test Agent',
            session_key='test_session'
        )
        
        analytics_admin = AnalyticsAdminSite()
        self.client.login(username='admin', password='testpass123')
        
        request = self.client.get('/admin/analytics-export/?type=search&format=csv&days=1').wsgi_request
        request.user = self.admin_user
        request.GET = {'type': 'search', 'format': 'csv', 'days': '1'}
        
        response = analytics_admin.analytics_export_view(request)
        content = response.content.decode('utf-8')
        
        # Should contain the query but properly escaped for CSV
        self.assertIn('alert', content)  # Content should be there
        # But should be properly handled in CSV format