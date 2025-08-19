"""
Data export and import utilities for API endpoints.
"""

import csv
import json
import xml.etree.ElementTree as ET
from io import StringIO, BytesIO
from django.http import HttpResponse, JsonResponse
from django.core.serializers import serialize
from django.core.exceptions import ValidationError
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
from rest_framework.parsers import MultiPartParser, FormParser
from django.db import transaction
from django.utils import timezone
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import logging

logger = logging.getLogger(__name__)


class DataExportMixin:
    """Mixin to add data export capabilities to ViewSets."""
    
    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        """Export data as CSV."""
        try:
            queryset = self.filter_queryset(self.get_queryset())
            
            # Get field names for CSV headers
            fields = self.get_export_fields()
            
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{self.get_export_filename()}.csv"'
            
            writer = csv.writer(response)
            
            # Write headers
            writer.writerow(fields)
            
            # Write data rows
            for obj in queryset:
                row = []
                for field in fields:
                    value = self.get_field_value(obj, field)
                    row.append(str(value) if value is not None else '')
                writer.writerow(row)
            
            return response
        
        except Exception as e:
            logger.error(f"CSV export error: {str(e)}")
            return Response(
                {'error': 'Export failed'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def export_json(self, request):
        """Export data as JSON."""
        try:
            queryset = self.filter_queryset(self.get_queryset())
            serializer = self.get_serializer(queryset, many=True)
            
            response = JsonResponse(serializer.data, safe=False)
            response['Content-Disposition'] = f'attachment; filename="{self.get_export_filename()}.json"'
            
            return response
        
        except Exception as e:
            logger.error(f"JSON export error: {str(e)}")
            return Response(
                {'error': 'Export failed'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def export_xml(self, request):
        """Export data as XML."""
        try:
            queryset = self.filter_queryset(self.get_queryset())
            
            # Create XML structure
            root = ET.Element('data')
            model_name = queryset.model.__name__.lower()
            
            for obj in queryset:
                item = ET.SubElement(root, model_name)
                
                for field in self.get_export_fields():
                    field_elem = ET.SubElement(item, field)
                    value = self.get_field_value(obj, field)
                    field_elem.text = str(value) if value is not None else ''
            
            # Convert to string
            xml_str = ET.tostring(root, encoding='unicode')
            
            response = HttpResponse(xml_str, content_type='application/xml')
            response['Content-Disposition'] = f'attachment; filename="{self.get_export_filename()}.xml"'
            
            return response
        
        except Exception as e:
            logger.error(f"XML export error: {str(e)}")
            return Response(
                {'error': 'Export failed'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Export data as Excel file."""
        try:
            queryset = self.filter_queryset(self.get_queryset())
            fields = self.get_export_fields()
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self.get_export_filename()
            
            # Style for headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center")
            
            # Write headers
            for col, field in enumerate(fields, 1):
                cell = ws.cell(row=1, column=col, value=field.replace('_', ' ').title())
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Write data
            for row, obj in enumerate(queryset, 2):
                for col, field in enumerate(fields, 1):
                    value = self.get_field_value(obj, field)
                    ws.cell(row=row, column=col, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{self.get_export_filename()}.xlsx"'
            
            return response
        
        except Exception as e:
            logger.error(f"Excel export error: {str(e)}")
            return Response(
                {'error': 'Export failed'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def get_export_fields(self):
        """Get fields to include in export."""
        # Default implementation - override in subclasses
        model = self.queryset.model
        return [field.name for field in model._meta.fields if not field.name.endswith('_ptr')]
    
    def get_field_value(self, obj, field_name):
        """Get field value for export."""
        try:
            # Handle nested field access (e.g., 'author.username')
            if '.' in field_name:
                parts = field_name.split('.')
                value = obj
                for part in parts:
                    value = getattr(value, part, None)
                    if value is None:
                        break
                return value
            else:
                return getattr(obj, field_name, None)
        except Exception:
            return None
    
    def get_export_filename(self):
        """Get filename for export."""
        model_name = self.queryset.model.__name__.lower()
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        return f"{model_name}_export_{timestamp}"


class DataImportMixin:
    """Mixin to add data import capabilities to ViewSets."""
    
    parser_classes = [MultiPartParser, FormParser]
    
    @action(detail=False, methods=['post'])
    def import_csv(self, request):
        """Import data from CSV file."""
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            file_obj = request.FILES['file']
            
            # Read CSV data
            csv_data = file_obj.read().decode('utf-8')
            csv_reader = csv.DictReader(StringIO(csv_data))
            
            return self.process_import_data(list(csv_reader), 'csv')
        
        except Exception as e:
            logger.error(f"CSV import error: {str(e)}")
            return Response(
                {'error': f'Import failed: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['post'])
    def import_json(self, request):
        """Import data from JSON file."""
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            file_obj = request.FILES['file']
            json_data = json.load(file_obj)
            
            if not isinstance(json_data, list):
                return Response(
                    {'error': 'JSON file must contain an array of objects'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            return self.process_import_data(json_data, 'json')
        
        except json.JSONDecodeError as e:
            return Response(
                {'error': f'Invalid JSON: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.error(f"JSON import error: {str(e)}")
            return Response(
                {'error': f'Import failed: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        """Import data from Excel file."""
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            file_obj = request.FILES['file']
            
            # Read Excel data
            wb = openpyxl.load_workbook(file_obj)
            ws = wb.active
            
            # Get headers from first row
            headers = [cell.value for cell in ws[1]]
            
            # Read data rows
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))
                # Remove None values
                row_data = {k: v for k, v in row_data.items() if v is not None}
                if row_data:  # Only add non-empty rows
                    data.append(row_data)
            
            return self.process_import_data(data, 'excel')
        
        except Exception as e:
            logger.error(f"Excel import error: {str(e)}")
            return Response(
                {'error': f'Import failed: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    def process_import_data(self, data, file_format):
        """Process imported data."""
        if not data:
            return Response(
                {'error': 'No data found in file'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validate data size
        if len(data) > 1000:  # Limit import size
            return Response(
                {'error': 'Maximum 1000 records allowed per import'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Transform data for serializer
        transformed_data = []
        errors = []
        
        for i, row in enumerate(data):
            try:
                transformed_row = self.transform_import_row(row, file_format)
                if transformed_row:
                    transformed_data.append(transformed_row)
            except Exception as e:
                errors.append(f"Row {i+1}: {str(e)}")
        
        if errors and not transformed_data:
            return Response(
                {'errors': errors},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validate and save data
        try:
            with transaction.atomic():
                serializer = self.get_serializer(data=transformed_data, many=True)
                
                if serializer.is_valid():
                    instances = serializer.save()
                    
                    response_data = {
                        'imported_count': len(instances),
                        'message': f'Successfully imported {len(instances)} records'
                    }
                    
                    if errors:
                        response_data['warnings'] = errors
                        response_data['warning_count'] = len(errors)
                    
                    return Response(response_data, status=status.HTTP_201_CREATED)
                else:
                    return Response(
                        {'validation_errors': serializer.errors},
                        status=status.HTTP_400_BAD_REQUEST
                    )
        
        except Exception as e:
            logger.error(f"Import processing error: {str(e)}")
            return Response(
                {'error': f'Failed to save data: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def transform_import_row(self, row, file_format):
        """Transform import row data."""
        # Default implementation - override in subclasses
        return row


class AdvancedExportMixin:
    """Advanced export functionality with filtering and customization."""
    
    @action(detail=False, methods=['post'])
    def custom_export(self, request):
        """Custom export with user-defined fields and filters."""
        try:
            export_config = request.data
            
            # Validate export configuration
            if not isinstance(export_config, dict):
                return Response(
                    {'error': 'Invalid export configuration'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Get export format
            export_format = export_config.get('format', 'csv').lower()
            if export_format not in ['csv', 'json', 'xml', 'excel']:
                return Response(
                    {'error': 'Unsupported export format'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Apply custom filters
            queryset = self.get_custom_export_queryset(export_config)
            
            # Get custom fields
            fields = export_config.get('fields', self.get_export_fields())
            
            # Generate export
            if export_format == 'csv':
                return self.generate_custom_csv(queryset, fields, export_config)
            elif export_format == 'json':
                return self.generate_custom_json(queryset, fields, export_config)
            elif export_format == 'xml':
                return self.generate_custom_xml(queryset, fields, export_config)
            elif export_format == 'excel':
                return self.generate_custom_excel(queryset, fields, export_config)
        
        except Exception as e:
            logger.error(f"Custom export error: {str(e)}")
            return Response(
                {'error': 'Export failed'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def get_custom_export_queryset(self, config):
        """Get queryset with custom filters applied."""
        queryset = self.filter_queryset(self.get_queryset())
        
        # Apply date range filter
        date_from = config.get('date_from')
        date_to = config.get('date_to')
        
        if date_from:
            try:
                date_from = datetime.fromisoformat(date_from)
                queryset = queryset.filter(created_at__gte=date_from)
            except ValueError:
                pass
        
        if date_to:
            try:
                date_to = datetime.fromisoformat(date_to)
                queryset = queryset.filter(created_at__lte=date_to)
            except ValueError:
                pass
        
        # Apply custom filters
        custom_filters = config.get('filters', {})
        if custom_filters:
            queryset = queryset.filter(**custom_filters)
        
        # Apply ordering
        order_by = config.get('order_by')
        if order_by:
            queryset = queryset.order_by(order_by)
        
        # Apply limit
        limit = config.get('limit')
        if limit and isinstance(limit, int) and limit > 0:
            queryset = queryset[:min(limit, 10000)]  # Max 10k records
        
        return queryset
    
    def generate_custom_csv(self, queryset, fields, config):
        """Generate custom CSV export."""
        response = HttpResponse(content_type='text/csv')
        filename = config.get('filename', self.get_export_filename())
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        
        writer = csv.writer(response)
        
        # Write headers
        headers = [field.replace('_', ' ').title() for field in fields]
        writer.writerow(headers)
        
        # Write data
        for obj in queryset:
            row = [self.get_field_value(obj, field) for field in fields]
            writer.writerow(row)
        
        return response
    
    def generate_custom_json(self, queryset, fields, config):
        """Generate custom JSON export."""
        data = []
        
        for obj in queryset:
            item = {}
            for field in fields:
                item[field] = self.get_field_value(obj, field)
            data.append(item)
        
        response = JsonResponse(data, safe=False)
        filename = config.get('filename', self.get_export_filename())
        response['Content-Disposition'] = f'attachment; filename="{filename}.json"'
        
        return response
    
    def generate_custom_excel(self, queryset, fields, config):
        """Generate custom Excel export."""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        filename = config.get('filename', self.get_export_filename())
        ws.title = filename[:31]  # Excel sheet name limit
        
        # Write headers with styling
        for col, field in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col, value=field.replace('_', ' ').title())
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Write data
        for row, obj in enumerate(queryset, 2):
            for col, field in enumerate(fields, 1):
                value = self.get_field_value(obj, field)
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        
        return response


class ImportValidationMixin:
    """Mixin for advanced import validation."""
    
    def validate_import_data(self, data, file_format):
        """Validate import data before processing."""
        errors = []
        warnings = []
        
        # Check required fields
        required_fields = self.get_required_import_fields()
        
        for i, row in enumerate(data):
            row_errors = []
            
            # Check required fields
            for field in required_fields:
                if field not in row or not row[field]:
                    row_errors.append(f"Missing required field: {field}")
            
            # Custom validation
            custom_errors = self.validate_import_row(row, i)
            if custom_errors:
                row_errors.extend(custom_errors)
            
            if row_errors:
                errors.append(f"Row {i+1}: {'; '.join(row_errors)}")
        
        return {
            'valid': len(errors) == 0,
            'errors': errors,
            'warnings': warnings
        }
    
    def get_required_import_fields(self):
        """Get required fields for import."""
        # Override in subclasses
        return []
    
    def validate_import_row(self, row, row_index):
        """Validate individual import row."""
        # Override in subclasses
        return []